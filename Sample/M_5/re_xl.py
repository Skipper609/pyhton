import openpyxl 
wb = openpyxl.load_workbook("Student.xlsx")
sheet = wb.active

for i in range(1,sheet.max_row+1):
    for j in range(1,sheet.max_column+1):
        cell_obj = sheet.cell(row = i, column = j) 
        print(cell_obj.value, end = " ") 
    print()